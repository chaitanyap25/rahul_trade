from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import Workbook, load_workbook
from datetime import date
import requests
import time

class Login_OpstraPage:
    username = 'username'
    psw = 'password'
    login = 'kc-login'
    con = "//*[@id='app']/div[2]/main/div/div/div/div/div[3]/div/main/div/section[1]/div/div[3]/div/a[2]/div"
    lo = "//*[@id='app']/div[36]/div/div/div[3]/button/div"
    po = "//*[@id='app']/div[8]/nav/div/div[3]/a[2]/div/i"
    expand = "(//input[@type='checkbox'])[3]"

    def __init__(self, driver):
        self.driver = driver

# //-------------------------------------------

    def selectexpiry(self, exp_date):
        # expiry = "//*[@id='expirySelect']/option[@value='"+exp_date+"']"
        expiry = "// *[ @ id = 'menu-'] / div[3] / ul / li[ @ data - value = '"+exp_date+"']"
        print(expiry)
        self.driver.find_element(By.ID, self.expiry).click()

    def getoptiondata(self, stk_price):

        call_buy = 290
        put_buy = 200

        fut = "//*[@id = 'app']/div/div[3]/div[2]/div[2]/div/div/div[2]/div[1]/div[1]/div/div[1]"
        nifty_fut = self.driver.find_element(By.XPATH, fut).text
        fut = nifty_fut.split(" ")
        print(fut)
        total_row_even = "//*[@id='oc-table-body']/div/div[@class='rt-tr -even']"
        total_row_odd = "//*[@id='oc-table-body']/div/div[@class='rt-tr -odd']"
        cnt = len(total_row_even) + len(total_row_odd)

        print(cnt)
        # i = 0
        for i in range(2, int(cnt)-1):
            print("Row: "+str(i))
            print("//*[@id='oc-table-body']/div["+str(i)+"]/div/div[4]")
            nifty = self.driver.find_element(By.XPATH, "//*[@id='oc-table-body']/div["+str(i)+"]/div/div[4]").text
            print("Nifty : " + nifty)
            if nifty == stk_price:
                print("pass")
                stk = "//*[@id='oc-table-body']/div["+str(i)+"]/div/div[3]/div/div/div"
                call_ltp = self.driver.find_element(By.XPATH, stk).text
                print(call_ltp)
                put = "// *[@id = 'oc-table-body']/div["+str(i)+"]/div/div[6]/div/div/div"
                put_ltp = self.driver.find_element(By.XPATH, put).text
                print(put_ltp)
                break


        # write in to excel
        data = (
            (nifty_fut, call_ltp, put_ltp, date.today()),
        )
        file_path = 'TestData/Option_trade.xlsx'
        # wbb = Workbook()
        wb = load_workbook(file_path)
        sheet = wb['Sheet']
        # sheet = wbb.active

        # wb = xw.Book('TestData/Option_trade.xlsx')
        # ws = wb.sheets["Sheet"]
        # a_range = ws.used_range.address
        print(sheet.max_row)
        new_row = sheet.max_row + 1
        ce = call_ltp.split(".")
        pe = put_ltp.split(".")
        print(call_ltp[0:6])
        profit = (call_buy - float(call_ltp[0:6])) + (put_buy - float(put_ltp[0:6]))
        total = profit*50
        print(total)

        named_tuple = time.localtime()  # get struct_time
        time_string = time.strftime("%H:%M:%S", named_tuple)

        # for i in data:
            # sheet.append(sheet.max_row+1)

        sheet.cell(row=new_row, column=1).value = nifty_fut
        sheet.cell(row=new_row, column=2).value = str(call_ltp[0:6])
        sheet.cell(row=new_row, column=3).value = str(put_ltp[0:6])
        sheet.cell(row=new_row, column=4).value = date.today()
        sheet.cell(row=new_row, column=5).value = time_string
        sheet.cell(row=new_row, column=6).value = round(total,2)

        wb.save('TestData/Option_trade.xlsx')
        wb.close()

        # ------------------------send to telegram----------------------------------------------
        msg = "Chaitanya : Your profit is "+str(round(total,2)) + " on time " +str(time_string)+" with CE at "+str(ce[0])+" and PE at "+str(pe[0])
        url1 = 'https://api.telegram.org/bot6006884871:AAFqjs2rjTKfn7LYonjdmogq6v4-LAEegTU/sendMessage?chat_id=-894738745&text="{}"'.format(msg)
#         requests.get(url)
        self.driver.get(url1)
